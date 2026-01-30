"""Enhanced unified dashboard - Primary entry point for the attendance system."""
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from datetime import datetime
from pathlib import Path
import pandas as pd

from config import (
    ADMIN_USERNAME, ADMIN_PASSWORD, STUDENT_CSV, ATTENDANCE_DIR,
    TRAINING_DIR, LABEL_DIR, ensure_data_dirs
)
from components.student_registration import register_student
from components.model_training import train_model
from components.auto_attendance import mark_auto_attendance
from components.manual_attendance import mark_manual_attendance
from components.analytics import compute_summary, daily_counts
from components.dashboard import Dashboard
from data.database_handler import read_students, list_csv_files
from utils.logger import log_info, log_error


class EnhancedDashboard(tk.Tk):
    """Unified dashboard with admin features, quick actions, and real-time statistics."""
    
    def __init__(self):
        super().__init__()
        self.title("📊 Attendance Management System - Enhanced Dashboard")
        self.geometry("1400x850")
        self.configure(bg="#0a1e3f")
        
        ensure_data_dirs()
        self.admin_logged_in = False
        self.current_user = None
        
        # Maximize window
        self.state('zoomed')
        
        self._build_ui()
        log_info("Dashboard launched")
    
    def on_manage_students(self):
        """Open student management window (admin only)."""
        if not self.admin_logged_in:
            messagebox.showwarning("Access Denied", "Only admin can manage students. Please login as admin first.")
            return
        self._open_student_management()

    def _open_student_management(self):
        """Open window to view, edit, and delete students."""
        mgmt_win = tk.Toplevel(self)
        mgmt_win.title("📊 Student Management")
        mgmt_win.geometry("1200x700")
        mgmt_win.minsize(900, 500)
        mgmt_win.configure(bg="#0a1e3f")
        mgmt_win.transient(self)
        
        # Header
        header = tk.Frame(mgmt_win, bg="#1a3a63", height=90)
        header.pack(fill="x", pady=0)
        header.pack_propagate(False)
        
        title_frame = tk.Frame(header, bg="#1a3a63")
        title_frame.pack(fill="x", padx=20, pady=15)
        
        tk.Label(title_frame, text="👥 Registered Students Management", bg="#1a3a63", fg="#00d4ff",
                font=("Arial", 18, "bold")).pack(anchor="w")
        
        tk.Label(title_frame, text="Manage student registrations, edit details, or remove records", 
                bg="#1a3a63", fg="#b0b0b0", font=("Arial", 9)).pack(anchor="w", pady=(3, 0))
        
        # Button frame at the bottom - create BEFORE content for proper layout
        btn_frame = tk.Frame(mgmt_win, bg="#0f2342", height=60)
        btn_frame.pack(fill="x", side="bottom", padx=0, pady=0)
        btn_frame.pack_propagate(False)
        
        # Content frame - using pack for proper expansion
        content = tk.Frame(mgmt_win, bg="#0a1e3f")
        content.pack(fill="both", expand=True, padx=15, pady=(15, 0))
        
        # Search frame with improved styling
        search_frame = tk.Frame(content, bg="#1a3a63", relief="flat", bd=1)
        search_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(search_frame, text="🔍 Search Students:", bg="#1a3a63", fg="#00d4ff",
                font=("Arial", 11, "bold")).pack(side="left", padx=12, pady=12)
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Arial", 11),
                               relief="flat", bd=0, bg="#0f2342", fg="white", insertbackground="white")
        search_entry.pack(side="left", fill="x", expand=True, padx=(0, 12), pady=12)
        search_entry.insert(0, "Search by name or enrollment ID...")
        
        def on_focus_in(event):
            if search_entry.get() == "Search by name or enrollment ID...":
                search_entry.delete(0, tk.END)
                search_entry.config(fg="white")
        
        def on_focus_out(event):
            if search_entry.get() == "":
                search_entry.insert(0, "Search by name or enrollment ID...")
                search_entry.config(fg="#555555")
        
        search_entry.bind("<FocusIn>", on_focus_in)
        search_entry.bind("<FocusOut>", on_focus_out)
        
        # Stats frame
        stats_frame = tk.Frame(content, bg="#1a3a63")
        stats_frame.pack(fill="x", pady=(0, 10))
        
        total_label = tk.Label(stats_frame, text="Total Students: 0", bg="#1a3a63", fg="#00d4ff",
                              font=("Arial", 10, "bold"))
        total_label.pack(anchor="w", padx=12, pady=8)
        
        # Treeview for students with better columns
        columns = ("Enrollment", "Name", "Date", "Time")
        tree = tk.ttk.Treeview(content, columns=columns, height=15, show="headings")
        
        tree.column("Enrollment", width=120, anchor="center", stretch=True)
        tree.column("Name", width=300, anchor="w", stretch=True)
        tree.column("Date", width=150, anchor="center", stretch=True)
        tree.column("Time", width=150, anchor="center", stretch=True)
        
        tree.heading("Enrollment", text="📋 Enrollment ID")
        tree.heading("Name", text="👤 Student Name")
        tree.heading("Date", text="📅 Registration Date")
        tree.heading("Time", text="⏰ Registration Time")
        
        # Configure style for better appearance
        style = tk.ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", background="#1a2a3f", foreground="white",
                       fieldbackground="#1a2a3f", font=("Arial", 10), rowheight=35, relief="flat")
        style.configure("Treeview.Heading", background="#366092", foreground="#00d4ff",
                       font=("Arial", 11, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#0066cc")], foreground=[("selected", "white")])
        
        # Tree frame with scrollbars
        tree_frame = tk.Frame(content, bg="#0a1e3f")
        tree_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Vertical scrollbar
        vsb = tk.ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.pack(side="right", fill="y")
        
        # Horizontal scrollbar
        hsb = tk.ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side="bottom", fill="x")
        
        tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        tree.pack(side="left", fill="both", expand=True)
        
        def load_students(search_text=""):
            """Load students into treeview."""
            for item in tree.get_children():
                tree.delete(item)
            
            try:
                df = read_students()
                if df.empty:
                    total_label.config(text="Total Students: 0")
                    tree.insert("", tk.END, values=("No students", "registered", "", ""))
                    return
                
                # Filter based on search text
                display_count = 0
                for _, row in df.iterrows():
                    if search_text.lower() in str(row.get('Name', '')).lower() or \
                       search_text.lower() in str(row.get('Enrollment', '')).lower():
                        tree.insert("", tk.END, values=(
                            row.get('Enrollment', ''),
                            row.get('Name', ''),
                            row.get('Date', ''),
                            row.get('Time', '')
                        ))
                        display_count += 1
                
                total_label.config(text=f"Total Students: {len(df)} (Showing: {display_count})")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load students: {str(e)}")
                total_label.config(text="Total Students: Error")
        
        def on_search_change(*args):
            """Handle search input change."""
            search_text = search_var.get()
            if search_text == "Search by name or enrollment ID...":
                load_students("")
            else:
                load_students(search_text)
        
        search_var.trace("w", on_search_change)
        load_students()
        
        def edit_student():
            """Edit selected student."""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Select Student", "Please select a student to edit.")
                return
            
            selected_item = selection[0]
            values = tree.item(selected_item, "values")
            old_enrollment, old_name, old_date, old_time = values
            
            # Edit window
            edit_win = tk.Toplevel(mgmt_win)
            edit_win.title(f"Edit Student - {old_enrollment}")
            edit_win.geometry("500x450")
            edit_win.minsize(400, 350)
            edit_win.configure(bg="#0a1e3f")
            edit_win.transient(mgmt_win)
            edit_win.grab_set()
            edit_win.grid_rowconfigure(1, weight=1)
            edit_win.grid_columnconfigure(0, weight=1)
            
            # Header
            header_edit = tk.Frame(edit_win, bg="#1a3a63")
            header_edit.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
            
            tk.Label(header_edit, text="✏️ Edit Student Information", bg="#1a3a63", fg="#00d4ff",
                    font=("Arial", 14, "bold"), pady=15).pack(fill="x", padx=20)
            
            form = tk.Frame(edit_win, bg="#0a1e3f")
            form.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)
            form.grid_columnconfigure(1, weight=1)
            
            # Enrollment ID field
            tk.Label(form, text="Enrollment ID:", bg="#0a1e3f", fg="#00d4ff",
                    font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 5))
            enr_entry = tk.Entry(form, font=("Arial", 11), relief="flat", bd=0, 
                                bg="#1a3a63", fg="white", insertbackground="white")
            enr_entry.grid(row=0, column=1, sticky="ew", pady=(0, 20))
            enr_entry.insert(0, old_enrollment)
            
            # Student Name field
            tk.Label(form, text="Student Name:", bg="#0a1e3f", fg="#00d4ff",
                    font=("Arial", 11, "bold")).grid(row=1, column=0, sticky="w", pady=(0, 5))
            name_entry = tk.Entry(form, font=("Arial", 11), relief="flat", bd=0,
                                 bg="#1a3a63", fg="white", insertbackground="white")
            name_entry.grid(row=1, column=1, sticky="ew", pady=(0, 30))
            name_entry.insert(0, old_name)
            name_entry.focus()
            
            # Button frame at bottom
            btn_frame_edit = tk.Frame(form, bg="#0a1e3f")
            btn_frame_edit.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(20, 0))
            btn_frame_edit.grid_columnconfigure(0, weight=1)
            btn_frame_edit.grid_columnconfigure(1, weight=1)
            
            def save_changes():
                """Save edited student info."""
                new_enrollment = enr_entry.get().strip()
                new_name = name_entry.get().strip()
                
                if not new_enrollment:
                    messagebox.showerror("Invalid", "Enrollment ID cannot be empty.")
                    return
                
                if not new_name:
                    messagebox.showerror("Invalid", "Student name cannot be empty.")
                    return
                
                # Check if new enrollment ID already exists
                if new_enrollment != old_enrollment:
                    df = read_students()
                    existing_ids = df['Enrollment'].astype(str).tolist()
                    if new_enrollment in existing_ids:
                        messagebox.showerror("Duplicate", f"Enrollment ID '{new_enrollment}' already exists!")
                        return
                
                try:
                    df = read_students()
                    mask = df['Enrollment'].astype(str) == str(old_enrollment)
                    df.loc[mask, 'Enrollment'] = new_enrollment
                    df.loc[mask, 'Name'] = new_name
                    df.to_csv(STUDENT_CSV, index=False)
                    messagebox.showinfo("Success", f"✅ Student updated successfully!\nEnrollment: {new_enrollment}\nName: {new_name}")
                    log_info(f"Student updated: {old_enrollment} -> {new_enrollment}, Name: {new_name}")
                    edit_win.destroy()
                    load_students(search_var.get())
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update student: {str(e)}")
                    log_error(f"Edit student error: {str(e)}")
            
            tk.Button(btn_frame_edit, text="💾 Save Changes", command=save_changes,
                     bg="#28a745", fg="white", font=("Arial", 11, "bold"),
                     relief="flat", bd=0, cursor="hand2", pady=12,
                     activebackground="#20c997", activeforeground="white").grid(row=0, column=0, sticky="ew", padx=(0, 5))
            
            tk.Button(btn_frame_edit, text="❌ Cancel", command=edit_win.destroy,
                     bg="#6c757d", fg="white", font=("Arial", 11, "bold"),
                     relief="flat", bd=0, cursor="hand2", pady=12,
                     activebackground="#5a6268", activeforeground="white").grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        def delete_student():
            """Delete selected student."""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Select Student", "Please select a student to delete.")
                return
            
            selected_item = selection[0]
            values = tree.item(selected_item, "values")
            enrollment_id, name = values[0], values[1]
            
            # Confirm deletion
            if not messagebox.askyesno("Confirm Delete", 
                                      f"Are you sure you want to delete?\n\n📋 Enrollment ID: {enrollment_id}\n👤 Name: {name}\n\n⚠️ This action cannot be undone!"):
                return
            
            try:
                df = read_students()
                initial_count = len(df)
                df = df[df['Enrollment'].astype(str) != str(enrollment_id)]
                
                if len(df) < initial_count:
                    df.to_csv(STUDENT_CSV, index=False)
                    messagebox.showinfo("Success", f"✅ Student '{name}' ({enrollment_id}) has been deleted successfully.")
                    log_info(f"Student deleted: {enrollment_id} - {name}")
                    load_students(search_var.get())
                else:
                    messagebox.showwarning("Not Found", f"Could not find student with ID: {enrollment_id}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete student: {str(e)}")
                log_error(f"Delete student error: {str(e)}")
        
        # Add buttons to the button frame created earlier
        tk.Button(btn_frame, text="✏️ Edit Student", command=edit_student,
                 bg="#0d6efd", fg="white", font=("Arial", 11, "bold"),
                 relief="flat", bd=0, cursor="hand2",
                 activebackground="#0b5ed7", activeforeground="white").pack(side="left", fill="both", expand=True)
        
        tk.Button(btn_frame, text="🗑️ Delete Student", command=delete_student,
                 bg="#dc3545", fg="white", font=("Arial", 11, "bold"),
                 relief="flat", bd=0, cursor="hand2",
                 activebackground="#bb2d3b", activeforeground="white").pack(side="left", fill="both", expand=True)
        
        tk.Button(btn_frame, text="🔄 Refresh", command=lambda: load_students(search_var.get()),
                 bg="#17a2b8", fg="white", font=("Arial", 11, "bold"),
                 relief="flat", bd=0, cursor="hand2",
                 activebackground="#138496", activeforeground="white").pack(side="left", fill="both", expand=True)
        
        tk.Button(btn_frame, text="❌ Close", command=mgmt_win.destroy,
                 bg="#6c757d", fg="white", font=("Arial", 11, "bold"),
                 relief="flat", bd=0, cursor="hand2",
                 activebackground="#5a6268", activeforeground="white").pack(side="left", fill="both", expand=True)

    def _build_ui(self):
        """Build the main dashboard UI."""
        # Header
        header = tk.Frame(self, bg="#1a3a63", height=80)
        header.pack(fill="x", pady=0)
        header.pack_propagate(False)
        
        title = tk.Label(header, text="🎯 Attendance Management Dashboard", 
                        bg="#1a3a63", fg="white", font=("Arial", 32, "bold"))
        title.pack(side="left", padx=30, pady=15)
        
        admin_btn = tk.Button(header, text="👤 Admin Login", 
                             command=self._admin_login, bg="#dc3545", fg="white",
                             font=("Arial", 11, "bold"), relief="raised", bd=2,
                             cursor="hand2", padx=15, pady=8)
        admin_btn.pack(side="right", padx=20, pady=15)
        self.admin_btn = admin_btn
        
        # Main container
        main_frame = tk.Frame(self, bg="#0a1e3f")
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Left side - Quick actions
        left_frame = tk.Frame(main_frame, bg="#0a1e3f", width=280)
        left_frame.pack(side="left", fill="both", padx=(0, 10))
        left_frame.pack_propagate(False)
        
        tk.Label(left_frame, text="⚡ Quick Actions", bg="#0a1e3f", fg="white",
                font=("Arial", 16, "bold"), pady=10).pack(fill="x")
        
        buttons_data = [
            ("👤 Register Student", self.on_register, "#007bff"),
            ("👥 Manage Students", self.on_manage_students, "#6c757d"),
            ("🧠 Train Model", self.on_train, "#28a745"),
            ("📹 Auto Attendance", self.on_auto_attend, "#17a2b8"),
            ("✏️ Manual Attendance", self.on_manual_attend, "#ffc107"),
            ("📊 View Analytics", self.on_analytics, "#6f42c1"),
            ("📥 Download Reports", self.on_download_reports, "#dc3545"),
        ]
        
        for btn_text, cmd, color in buttons_data:
            btn = tk.Button(left_frame, text=btn_text, command=cmd,
                          bg=color, fg="white", font=("Arial", 12, "bold"),
                          relief="raised", bd=2, cursor="hand2", width=25)
            btn.pack(pady=8, fill="x")
        
        # Right side - Statistics
        right_frame = tk.Frame(main_frame, bg="#0a1e3f")
        right_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        tk.Label(right_frame, text="📈 System Statistics", bg="#0a1e3f", fg="white",
                font=("Arial", 16, "bold"), pady=10).pack(fill="x")
        
        stats_container = tk.Frame(right_frame, bg="#0a1e3f")
        stats_container.pack(fill="both", expand=True)
        
        self._add_stat_card(stats_container, "👥 Total Students", 
                           self._get_student_count(), "#007bff")
        self._add_stat_card(stats_container, "📋 Total Records", 
                           self._get_record_count(), "#28a745")
        self._add_stat_card(stats_container, "✅ Today's Attendance", 
                           self._get_today_count(), "#17a2b8")
        
        # Recent activity
        activity_label = tk.Label(right_frame, text="📝 Recent Activity", bg="#0a1e3f",
                                 fg="white", font=("Arial", 14, "bold"), pady=10)
        activity_label.pack(fill="x", pady=(20, 5))
        
        activity_frame = tk.Frame(right_frame, bg="white", relief="sunken", bd=2)
        activity_frame.pack(fill="both", expand=True)
        
        self.activity_text = tk.Text(activity_frame, height=10, bg="white",
                                    font=("Arial", 10), relief="flat")
        self.activity_text.pack(fill="both", expand=True, padx=5, pady=5)
        self.activity_text.config(state="disabled")
        
        # Show latest records
        self._update_activity()
        
        # Footer
        footer = tk.Frame(self, bg="#1a3a63", height=40)
        footer.pack(fill="x", side="bottom", pady=0)
        footer.pack_propagate(False)
        
        footer_text = tk.Label(footer, 
                              text="Confidence Threshold: < 70 | Auto-restart: 2 seconds | Duplicate-free attendance",
                              bg="#1a3a63", fg="#a8d5ff", font=("Arial", 10))
        footer_text.pack(side="left", padx=15, pady=10)
        
        time_label = tk.Label(footer, text=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                             bg="#1a3a63", fg="#a8d5ff", font=("Arial", 10))
        time_label.pack(side="right", padx=15, pady=10)
        self.time_label = time_label
        
        self._update_time()
    
    def _add_stat_card(self, parent, label, value, color):
        """Add a statistics card."""
        card = tk.Frame(parent, bg=color, relief="raised", bd=2)
        card.pack(fill="x", pady=10)
        
        label_widget = tk.Label(card, text=label, bg=color, fg="white",
                               font=("Arial", 12, "bold"), justify="left", padx=15, pady=8)
        label_widget.pack(side="left")
        
        value_widget = tk.Label(card, text=str(value), bg=color, fg="white",
                               font=("Arial", 24, "bold"), padx=15, pady=8)
        value_widget.pack(side="right")
    
    def _get_student_count(self):
        """Get total registered students."""
        try:
            df = read_students()
            return len(df) if not df.empty else 0
        except:
            return 0
    
    def _get_record_count(self):
        """Get total attendance records."""
        try:
            total = 0
            for csv_file in ATTENDANCE_DIR.glob("*.csv"):
                try:
                    df = pd.read_csv(csv_file)
                    total += len(df)
                except:
                    pass
            return total
        except:
            return 0
    
    def _get_today_count(self):
        """Get today's attendance count."""
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            total = 0
            for csv_file in ATTENDANCE_DIR.glob("*.csv"):
                try:
                    df = pd.read_csv(csv_file)
                    today_records = df[df['Date'] == today]
                    total += len(today_records)
                except:
                    pass
            return total
        except:
            return 0
    
    def _update_activity(self):
        """Update recent activity display."""
        self.activity_text.config(state="normal")
        self.activity_text.delete("1.0", tk.END)
        
        try:
            # Get latest CSV files
            csv_files = sorted(list_csv_files(ATTENDANCE_DIR), reverse=True)[:5]
            
            if not csv_files:
                self.activity_text.insert(tk.END, "No recent activity")
            else:
                for csv_file in csv_files:
                    try:
                        df = pd.read_csv(ATTENDANCE_DIR / csv_file)
                        self.activity_text.insert(tk.END, f"📄 {csv_file}\n")
                        self.activity_text.insert(tk.END, f"   Records: {len(df)} students\n\n")
                    except:
                        pass
        except Exception as e:
            self.activity_text.insert(tk.END, f"Error loading activity: {str(e)}")
        
        self.activity_text.config(state="disabled")
    
    def _update_time(self):
        """Update time display and refresh stats."""
        try:
            if self.winfo_exists():
                self.time_label.config(text=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                self.after(1000, self._update_time)
        except tk.TclError:
            pass  # Window was closed, stop updating
    
    def _admin_login(self):
        """Admin login dialog."""
        login_win = tk.Toplevel(self)
        login_win.title("Admin Login")
        login_win.geometry("400x250")
        login_win.configure(bg="#1e3a5f")
        login_win.resizable(False, False)
        
        # Center on screen
        login_win.transient(self)
        login_win.grab_set()
        
        tk.Label(login_win, text="🔐 Admin Login", bg="#1e3a5f", fg="white",
                font=("Arial", 18, "bold"), pady=15).pack(fill="x")
        
        form_frame = tk.Frame(login_win, bg="#1e3a5f")
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        tk.Label(form_frame, text="Username:", bg="#1e3a5f", fg="white",
                font=("Arial", 11)).pack(anchor="w", pady=(0, 5))
        username_entry = tk.Entry(form_frame, font=("Arial", 11), relief="sunken", bd=2)
        username_entry.pack(fill="x", pady=(0, 15))
        username_entry.focus()
        
        tk.Label(form_frame, text="Password:", bg="#1e3a5f", fg="white",
                font=("Arial", 11)).pack(anchor="w", pady=(0, 5))
        password_entry = tk.Entry(form_frame, font=("Arial", 11), relief="sunken", bd=2, show="•")
        password_entry.pack(fill="x", pady=(0, 20))
        
        def login_check():
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            
            if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                self.admin_logged_in = True
                self.current_user = username
                self.admin_btn.config(text=f"👤 {username} (Logout)", command=self._admin_logout)
                messagebox.showinfo("Success", f"Welcome, {username}! Admin features enabled.")
                log_info(f"Admin login: {username}")
                login_win.destroy()
            else:
                messagebox.showerror("Failed", "Invalid credentials!")
        
        tk.Button(form_frame, text="🔓 Login", command=login_check,
                 bg="#28a745", fg="white", font=("Arial", 12, "bold"),
                 relief="raised", bd=2, cursor="hand2").pack(fill="x", pady=10)
        
        password_entry.bind("<Return>", lambda e: login_check())
    
    def _admin_logout(self):
        """Admin logout."""
        self.admin_logged_in = False
        self.current_user = None
        self.admin_btn.config(text="👤 Admin Login", command=self._admin_login)
        messagebox.showinfo("Logout", "Admin session ended.")
        log_info("Admin logout")
    
    def on_register(self):
        """Open student registration (admin only)."""
        if not self.admin_logged_in:
            messagebox.showwarning("Access Denied", "Only admin can register new students. Please login as admin first.")
            return
        
        try:
            register_student(self)
            self._update_activity()
        except Exception as e:
            messagebox.showerror("Error", f"Registration error: {str(e)}")
            log_info(f"Registration error: {str(e)}")
    
    def on_train(self):
        """Open model training."""
        try:
            train_model(self)
        except Exception as e:
            messagebox.showerror("Error", f"Training error: {str(e)}")
    
    def on_auto_attend(self):
        """Open auto attendance."""
        try:
            subject = simpledialog.askstring("Subject", "Enter subject/class name:", 
                                           parent=self)
            if subject:
                mark_auto_attendance(self, subject)
                self._update_activity()
        except Exception as e:
            messagebox.showerror("Error", f"Attendance error: {str(e)}")
    
    def on_manual_attend(self):
        """Open manual attendance."""
        try:
            mark_manual_attendance(self)
            self._update_activity()
        except Exception as e:
            messagebox.showerror("Error", f"Manual attendance error: {str(e)}")
    
    def on_analytics(self):
        """Open analytics dashboard."""
        try:
            Dashboard(self)
        except Exception as e:
            messagebox.showerror("Error", f"Analytics error: {str(e)}")
    
    def on_download_reports(self):
        """Download attendance reports (admin only)."""
        if not self.admin_logged_in:
            messagebox.showwarning("Access Denied", "Admin login required!")
            return
        
        # Ask user for format preference
        format_choice = messagebox.askyesnocancel(
            "Select Format",
            "Download as Excel format?\n\nYes = Excel (.xlsx)\nNo = CSV files\nCancel = Abort"
        )
        
        if format_choice is None:  # User clicked Cancel
            return
            
        folder = filedialog.askdirectory(title="Select folder to save reports")
        if not folder:
            return
        
        try:
            folder_path = Path(folder)
            
            if format_choice:  # Excel format
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    # Create a single Excel file with all attendance data
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)  # Remove default sheet
                    
                    all_data = []
                    subject_data = {}
                    total_files = 0

                    students_df = read_students()
                    if not students_df.empty and 'Enrollment' in students_df.columns and 'Name' in students_df.columns:
                        students_df['Enrollment'] = students_df['Enrollment'].astype(str)
                        name_map = dict(zip(students_df['Enrollment'], students_df['Name']))
                    else:
                        name_map = {}
                    
                    for csv_file in ATTENDANCE_DIR.glob("*.csv"):
                        try:
                            df = pd.read_csv(csv_file)
                            df['Session'] = csv_file.stem  # Add session info

                            if 'Enrollment' in df.columns:
                                df['Enrollment'] = df['Enrollment'].astype(str)
                                if 'Name' in df.columns:
                                    df['Name'] = df['Enrollment'].map(name_map).fillna(df['Name'])
                                else:
                                    df['Name'] = df['Enrollment'].map(name_map).fillna("")

                            all_data.append(df)

                            subject = csv_file.stem.split("_")[0] if "_" in csv_file.stem else "Unknown"
                            subject_data.setdefault(subject, []).append(df)
                            
                            # Create individual sheet for each session
                            ws = wb.create_sheet(title=csv_file.stem[:31])  # Excel limit 31 chars
                            
                            # Add title
                            ws.append([f"Attendance Report: {csv_file.stem}"])
                            ws['A1'].font = Font(bold=True, size=14)
                            ws.merge_cells('A1:D1')
                            
                            # Add headers
                            ws.append(['Enrollment', 'Name', 'Date', 'Time'])
                            for cell in ws[2]:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal='center')
                            
                            # Add data
                            for _, row in df.iterrows():
                                ws.append([row['Enrollment'], row['Name'], row['Date'], row['Time']])
                            
                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = None
                                for cell in column:
                                    try:
                                        if hasattr(cell, 'column_letter'):
                                            column_letter = cell.column_letter
                                        if hasattr(cell, 'value') and cell.value:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                    except:
                                        pass
                                if column_letter:
                                    adjusted_width = min(max_length + 2, 50)
                                    ws.column_dimensions[column_letter].width = adjusted_width
                            
                            total_files += 1
                        except Exception as e:
                            log_error(f"Error processing {csv_file.name}: {e}")
                    
                    # Create summary sheet
                    if all_data:
                        summary_df = pd.concat(all_data, ignore_index=True)
                        ws_summary = wb.create_sheet(title="All Records", index=0)
                        
                        ws_summary.append(["Complete Attendance Records"])
                        ws_summary['A1'].font = Font(bold=True, size=14)
                        ws_summary.merge_cells('A1:E1')
                        
                        ws_summary.append(['Enrollment', 'Name', 'Date', 'Time', 'Session'])
                        for cell in ws_summary[2]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center')
                        
                        for _, row in summary_df.iterrows():
                            ws_summary.append([row['Enrollment'], row['Name'], row['Date'], row['Time'], row['Session']])
                        
                        # Auto-adjust columns
                        for column in ws_summary.columns:
                            max_length = 0
                            column_letter = None
                            for cell in column:
                                try:
                                    if hasattr(cell, 'column_letter'):
                                        column_letter = cell.column_letter
                                    if hasattr(cell, 'value') and cell.value:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                except:
                                    pass
                            if column_letter:
                                adjusted_width = min(max_length + 2, 50)
                                ws_summary.column_dimensions[column_letter].width = adjusted_width

                    # Create subject-wise sheets
                    for subject, frames in subject_data.items():
                        subject_df = pd.concat(frames, ignore_index=True)
                        sheet_title = subject[:31]
                        if sheet_title in wb.sheetnames:
                            suffix = 1
                            while f"{sheet_title[:28]}_{suffix}" in wb.sheetnames:
                                suffix += 1
                            sheet_title = f"{sheet_title[:28]}_{suffix}"

                        ws_subject = wb.create_sheet(title=sheet_title)
                        ws_subject.append([f"Attendance Report: {subject}"])
                        ws_subject['A1'].font = Font(bold=True, size=14)
                        ws_subject.merge_cells('A1:E1')

                        ws_subject.append(['Enrollment', 'Name', 'Date', 'Time', 'Session'])
                        for cell in ws_subject[2]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center')

                        for _, row in subject_df.iterrows():
                            ws_subject.append([row['Enrollment'], row['Name'], row['Date'], row['Time'], row['Session']])

                        for column in ws_subject.columns:
                            max_length = 0
                            column_letter = None
                            for cell in column:
                                try:
                                    if hasattr(cell, 'column_letter'):
                                        column_letter = cell.column_letter
                                    if hasattr(cell, 'value') and cell.value:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                except:
                                    pass
                            if column_letter:
                                adjusted_width = min(max_length + 2, 50)
                                ws_subject.column_dimensions[column_letter].width = adjusted_width
                    
                    # Save Excel file
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    excel_path = folder_path / f"Attendance_Report_{timestamp}.xlsx"
                    wb.save(excel_path)

                    # Save per-subject Excel files
                    for subject, frames in subject_data.items():
                        subject_df = pd.concat(frames, ignore_index=True)
                        subject_safe = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in subject)
                        subject_path = folder_path / f"Attendance_{subject_safe}_{timestamp}.xlsx"

                        subject_wb = openpyxl.Workbook()
                        subject_ws = subject_wb.active
                        subject_ws.title = subject[:31]

                        subject_ws.append([f"Attendance Report: {subject}"])
                        subject_ws['A1'].font = Font(bold=True, size=14)
                        subject_ws.merge_cells('A1:E1')

                        subject_ws.append(['Enrollment', 'Name', 'Date', 'Time', 'Session'])
                        for cell in subject_ws[2]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center')

                        for _, row in subject_df.iterrows():
                            subject_ws.append([row['Enrollment'], row['Name'], row['Date'], row['Time'], row['Session']])

                        for column in subject_ws.columns:
                            max_length = 0
                            column_letter = None
                            for cell in column:
                                try:
                                    if hasattr(cell, 'column_letter'):
                                        column_letter = cell.column_letter
                                    if hasattr(cell, 'value') and cell.value:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                except:
                                    pass
                            if column_letter:
                                adjusted_width = min(max_length + 2, 50)
                                subject_ws.column_dimensions[column_letter].width = adjusted_width

                        subject_wb.save(subject_path)
                    
                    messagebox.showinfo(
                        "Success",
                        f"Excel report created!\n{total_files} sessions exported\nFile: {excel_path.name}\nSubject files: {len(subject_data)}"
                    )
                    log_info(f"Admin {self.current_user} downloaded Excel report with {total_files} sessions")
                    
                except ImportError:
                    messagebox.showerror("Missing Package", "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
                    return
            else:  # CSV format
                total_files = 0
                for csv_file in ATTENDANCE_DIR.glob("*.csv"):
                    try:
                        import shutil
                        subject = csv_file.stem.split("_")[0] if "_" in csv_file.stem else "Unknown"
                        subject_folder = folder_path / subject
                        subject_folder.mkdir(parents=True, exist_ok=True)
                        shutil.copy(csv_file, subject_folder / csv_file.name)
                        total_files += 1
                    except Exception as e:
                        messagebox.showerror("Copy Error", f"Could not copy {csv_file.name}: {str(e)}")
                
                messagebox.showinfo("Success", f"Downloaded {total_files} CSV files!")
                log_info(f"Admin {self.current_user} downloaded {total_files} CSV reports")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to download reports: {str(e)}")
            log_error(f"Report download error: {str(e)}")


if __name__ == "__main__":
    app = EnhancedDashboard()
    app.mainloop()
